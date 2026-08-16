"""
Excel bridge for the Projection Model dashboard.

Serves dashboard.html on http://127.0.0.1:8765 and exposes three endpoints the
dashboard calls:

    GET  /api/health   -> {"ok": true}                     (used to detect the bridge)
    GET  /api/model    -> Company Inputs read from the .xlsx (cached values)
    POST /api/refresh  -> writes the ticker into the workbook, triggers Excel's
                          Power Query refresh (RefreshAll) via COM, waits for the
                          queries to settle, saves, and returns the fresh values.

Refresh has a documented fallback chain, so it degrades instead of dying:

    1. Excel COM (pywin32 + Excel installed)  -> real Power Query refresh
    2. Alpha Vantage called directly here     -> same two endpoints the M code uses
    3. Cached values already in the workbook  -> whatever the last refresh left

Only openpyxl is required. pywin32 is optional (step 1); without it the bridge
still works via step 2.

Run:  python excel_bridge.py            (or double-click run_dashboard.bat)
"""

import base64
import hmac
import json
import os
import re
import secrets
import sys
import threading
import time
import urllib.parse
import urllib.request
import webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

import openpyxl

import notion_service
import quotes_service

HERE = os.path.dirname(os.path.abspath(__file__))
WORKBOOK = os.path.normpath(os.path.join(HERE, "..", "Financial_Projection_Model.xlsx"))
SHEET = "Projection Model"
# Cloud hosts (Render, etc.) assign a port via $PORT and expect the app to
# bind it — local runs never set this, so the local default is untouched.
PORT = int(os.environ.get("PORT", 8765))
IS_CLOUD = "PORT" in os.environ

# A Notion *internal* integration token (from notion.so/my-integrations) —
# single-user, so this is one shared secret like BRIDGE_USERNAME/PASSWORD
# below, not a per-account OAuth token. Never sent to the frontend; every
# Notion call happens here on the server. Empty string = feature not set up.
NOTION_TOKEN = os.environ.get("NOTION_TOKEN", "").strip()
CREDENTIALS_FILE = os.path.join(HERE, ".bridge_credentials.json")


# ── auth ─────────────────────────────────────────────────────────────────────
# Basic Auth in front of every route. This matters most once the bridge is
# reachable off this machine (LAN or a public tunnel like ngrok) — without it,
# anyone with the URL could read the model and trigger an Excel refresh.
# Credentials are generated once and cached locally so they survive restarts;
# delete .bridge_credentials.json to force a new password.
def load_or_create_credentials():
    # Cloud deploys set these explicitly (Render's dashboard env vars) so the
    # password is known up front instead of buried in a log line, and so it
    # survives a redeploy on a host with an ephemeral filesystem.
    env_user, env_pass = os.environ.get("BRIDGE_USERNAME"), os.environ.get("BRIDGE_PASSWORD")
    if env_user and env_pass:
        return env_user, env_pass
    if os.path.exists(CREDENTIALS_FILE):
        try:
            with open(CREDENTIALS_FILE, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if data.get("username") and data.get("password"):
                return data["username"], data["password"]
        except Exception:                                        # noqa: BLE001
            pass
    creds = {"username": "owner", "password": secrets.token_urlsafe(9)}
    try:
        with open(CREDENTIALS_FILE, "w", encoding="utf-8") as fh:
            json.dump(creds, fh)
    except Exception:                                             # noqa: BLE001
        pass
    return creds["username"], creds["password"]


AUTH_USER, AUTH_PASS = load_or_create_credentials()
AUTH_TOKEN = base64.b64encode(f"{AUTH_USER}:{AUTH_PASS}".encode()).decode()
AUTH_ENABLED = "--no-auth" not in sys.argv[1:]

# Cell addresses of the Company Inputs block. These mirror the workbook's
# defined names (Ticker / ApiKey / Co_Revenue / Co_NetIncome / Co_EPS /
# Co_Shares); we resolve by name where possible and fall back to these.
FALLBACK_CELLS = {
    "ticker": "C3",
    "apiKey": "C4",
    "revenue": "C5",
    "netIncome": "C6",
    "eps": "C7",
    "shares": "C8",
}
NAME_MAP = {
    "ticker": "Ticker",
    "apiKey": "ApiKey",
    "revenue": "Co_Revenue",
    "netIncome": "Co_NetIncome",
    "eps": "Co_EPS",
    "shares": "Co_Shares",
}


# ── workbook reading ────────────────────────────────────────────────────────
def resolve_cells(wb):
    """Map logical field -> A1 address, preferring the workbook's defined names."""
    out = dict(FALLBACK_CELLS)
    for field, name in NAME_MAP.items():
        dn = wb.defined_names.get(name)
        if dn is None:
            continue
        try:
            # attr_text looks like: 'Projection Model'!$C$5
            ref = dn.attr_text.split("!")[-1].replace("$", "")
            if ref:
                out[field] = ref
        except Exception:
            pass
    return out


# ── saved projections: file-backed, so OneDrive syncs them across machines ──
# This folder already lives inside a OneDrive-synced directory, so writing
# each saved projection as its own JSON file here is enough to get real
# cross-device sync for free — no server, no accounts, just files OneDrive
# was already going to sync. The dashboard's localStorage save stays the
# fast local copy; this is the layer that lets a second PC see it too.
SAVES_DIR = os.path.join(HERE, "saves")
PROJECT_ID_RE = re.compile(r"^[A-Za-z0-9_-]{1,80}$")

# On a cloud host the filesystem is wiped on every restart and redeploy, so the
# files above can't be the storage layer there. Set NEON_DATABASE_URL and
# projections go to Postgres instead, which is what actually makes them survive
# a restart and show up on your other devices. Left unset — the normal
# local/OneDrive setup — nothing changes and the files stay the storage.
#
# Deliberately NOT named DATABASE_URL: Render reserves that exact key for its
# own native "connect a Render Postgres" env-var link and silently drops a
# manually-entered value under that name — it never reaches the process (verified
# by dumping os.environ.keys() at boot: every other var we set showed up,
# DATABASE_URL alone did not, with an external/Neon database behind it).
DATABASE_URL = (os.environ.get("NEON_DATABASE_URL") or os.environ.get("DATABASE_URL") or "").strip()
USE_PG = bool(DATABASE_URL)
_pg_error = None


def _pg_connect():
    import psycopg
    return psycopg.connect(DATABASE_URL, connect_timeout=10)


def pg_init():
    """Create the tables if needed. Returns (ok, error_message)."""
    global _pg_error
    try:
        with _pg_connect() as conn, conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS projections (
                    id         TEXT PRIMARY KEY,
                    data       TEXT NOT NULL,
                    updated_at TIMESTAMPTZ NOT NULL DEFAULT now()
                )
            """)
            # Single-row table — this bridge has one shared login, not
            # per-user accounts, so there's exactly one Notion database
            # connected at a time. See notion_settings_get/set below.
            cur.execute("""
                CREATE TABLE IF NOT EXISTS notion_settings (
                    id            TEXT PRIMARY KEY DEFAULT 'default',
                    database_id   TEXT NOT NULL,
                    database_name TEXT NOT NULL,
                    updated_at    TIMESTAMPTZ NOT NULL DEFAULT now()
                )
            """)
            # Same single-row shape as notion_settings — one shared watchlist,
            # stored as a JSON array of tickers.
            cur.execute("""
                CREATE TABLE IF NOT EXISTS watchlist (
                    id         TEXT PRIMARY KEY DEFAULT 'default',
                    tickers    TEXT NOT NULL,
                    updated_at TIMESTAMPTZ NOT NULL DEFAULT now()
                )
            """)
            conn.commit()
        _pg_error = None
        return True, None
    except Exception as exc:                                     # noqa: BLE001
        _pg_error = str(exc)
        return False, _pg_error


def _project_path(project_id):
    if not PROJECT_ID_RE.match(project_id or ""):
        return None
    return os.path.join(SAVES_DIR, project_id + ".json")


def list_saved_projects():
    """Every saved projection, keyed by id — same shape as the client's map."""
    out = {}
    if USE_PG:
        try:
            with _pg_connect() as conn, conn.cursor() as cur:
                cur.execute("SELECT data FROM projections")
                for (raw,) in cur.fetchall():
                    try:
                        data = json.loads(raw)
                    except json.JSONDecodeError:
                        continue        # skip one bad row, don't fail the list
                    if isinstance(data, dict) and data.get("id"):
                        out[data["id"]] = data
        except Exception as exc:                                 # noqa: BLE001
            print(f"[bridge] Postgres read failed: {exc}", flush=True)
        return out

    if not os.path.isdir(SAVES_DIR):
        return out
    for name in os.listdir(SAVES_DIR):
        if not name.endswith(".json"):
            continue
        try:
            with open(os.path.join(SAVES_DIR, name), "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if isinstance(data, dict) and data.get("id"):
                out[data["id"]] = data
        except (OSError, json.JSONDecodeError):
            continue  # skip a corrupt file rather than fail the whole list
    return out


def write_saved_project(data):
    """Create or overwrite one saved projection. Returns (ok, error_message)."""
    if not isinstance(data, dict) or not data.get("id") or not data.get("name"):
        return False, "Project must include at least 'id' and 'name'."
    if not PROJECT_ID_RE.match(data["id"]):
        return False, "Invalid project id."

    if USE_PG:
        # Surface a failure rather than silently falling back to the ephemeral
        # disk — the browser still holds its own copy, so the user loses
        # nothing, but they do need to know sync didn't happen.
        try:
            with _pg_connect() as conn, conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO projections (id, data, updated_at)
                    VALUES (%s, %s, now())
                    ON CONFLICT (id) DO UPDATE
                        SET data = EXCLUDED.data, updated_at = now()
                """, (data["id"], json.dumps(data)))
                conn.commit()
            return True, None
        except Exception as exc:                                 # noqa: BLE001
            return False, f"Could not save to database: {exc}"

    path = _project_path(data["id"])
    if path is None:
        return False, "Invalid project id."
    os.makedirs(SAVES_DIR, exist_ok=True)
    try:
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(data, fh, indent=2)
    except OSError as exc:
        return False, f"Could not write project file: {exc}"
    return True, None


def delete_saved_project(project_id):
    if not PROJECT_ID_RE.match(project_id or ""):
        return False

    if USE_PG:
        try:
            with _pg_connect() as conn, conn.cursor() as cur:
                cur.execute("DELETE FROM projections WHERE id = %s", (project_id,))
                deleted = cur.rowcount > 0
                conn.commit()
            return deleted
        except Exception as exc:                                 # noqa: BLE001
            print(f"[bridge] Postgres delete failed: {exc}", flush=True)
            return False

    path = _project_path(project_id)
    if path is None or not os.path.isfile(path):
        return False
    try:
        os.remove(path)
        return True
    except OSError:
        return False


# ── Notion Research: which database this bridge syncs stock pages into ──────
# Single row, same file/Postgres split as saved projections above — a local
# JSON file when there's no NEON_DATABASE_URL, Postgres when there is.
NOTION_SETTINGS_FILE = os.path.join(HERE, "notion_settings.json")


def get_notion_settings():
    """Returns {"databaseId", "databaseName"} or None if nothing's selected yet."""
    if USE_PG:
        try:
            with _pg_connect() as conn, conn.cursor() as cur:
                cur.execute("SELECT database_id, database_name FROM notion_settings WHERE id = 'default'")
                row = cur.fetchone()
            if not row:
                return None
            return {"databaseId": row[0], "databaseName": row[1]}
        except Exception as exc:                                  # noqa: BLE001
            print(f"[bridge] Postgres read (notion_settings) failed: {exc}", flush=True)
            return None

    try:
        with open(NOTION_SETTINGS_FILE, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        if data.get("databaseId") and data.get("databaseName"):
            return {"databaseId": data["databaseId"], "databaseName": data["databaseName"]}
    except (OSError, json.JSONDecodeError):
        pass
    return None


def set_notion_settings(database_id, database_name):
    if USE_PG:
        try:
            with _pg_connect() as conn, conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO notion_settings (id, database_id, database_name, updated_at)
                    VALUES ('default', %s, %s, now())
                    ON CONFLICT (id) DO UPDATE
                        SET database_id = EXCLUDED.database_id,
                            database_name = EXCLUDED.database_name,
                            updated_at = now()
                """, (database_id, database_name))
                conn.commit()
            return True, None
        except Exception as exc:                                  # noqa: BLE001
            return False, f"Could not save to database: {exc}"

    try:
        with open(NOTION_SETTINGS_FILE, "w", encoding="utf-8") as fh:
            json.dump({"databaseId": database_id, "databaseName": database_name}, fh, indent=2)
        return True, None
    except OSError as exc:
        return False, f"Could not write settings file: {exc}"


# ── Watchlist: a short list of tickers for the header strip ─────────────────
# Same file/Postgres split as everything else above. One shared list — this
# bridge has one user, not per-user accounts.
WATCHLIST_FILE = os.path.join(HERE, "watchlist.json")
WATCHLIST_MAX = 12
TICKER_RE = re.compile(r"^[A-Z][A-Z0-9.\-]{0,9}$")


def get_watchlist():
    """Returns the saved ticker list (uppercase strings), in saved order."""
    if USE_PG:
        try:
            with _pg_connect() as conn, conn.cursor() as cur:
                cur.execute("SELECT tickers FROM watchlist WHERE id = 'default'")
                row = cur.fetchone()
            if not row:
                return []
            data = json.loads(row[0])
            return data if isinstance(data, list) else []
        except Exception as exc:                                  # noqa: BLE001
            print(f"[bridge] Postgres read (watchlist) failed: {exc}", flush=True)
            return []

    try:
        with open(WATCHLIST_FILE, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        if isinstance(data, list):
            return [t for t in data if isinstance(t, str)]
    except (OSError, json.JSONDecodeError):
        pass
    return []


def set_watchlist(tickers):
    """Validates, dedupes, and caps the list, then saves it. Returns (ok, tickers_or_error)."""
    clean = []
    for raw in tickers:
        t = str(raw or "").strip().upper()
        if t and TICKER_RE.match(t) and t not in clean:
            clean.append(t)
    clean = clean[:WATCHLIST_MAX]

    if USE_PG:
        try:
            with _pg_connect() as conn, conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO watchlist (id, tickers, updated_at)
                    VALUES ('default', %s, now())
                    ON CONFLICT (id) DO UPDATE
                        SET tickers = EXCLUDED.tickers, updated_at = now()
                """, (json.dumps(clean),))
                conn.commit()
            return True, clean
        except Exception as exc:                                  # noqa: BLE001
            return False, f"Could not save to database: {exc}"

    try:
        with open(WATCHLIST_FILE, "w", encoding="utf-8") as fh:
            json.dump(clean, fh, indent=2)
        return True, clean
    except OSError as exc:
        return False, f"Could not write watchlist file: {exc}"


def read_company():
    """Read the Company Inputs block from the workbook's cached values."""
    if not os.path.exists(WORKBOOK):
        return {"error": f"Workbook not found: {WORKBOOK}"}
    try:
        wb_f = openpyxl.load_workbook(WORKBOOK)                  # for defined names
        cells = resolve_cells(wb_f)
        wb_f.close()
        wb = openpyxl.load_workbook(WORKBOOK, data_only=True)    # cached values
        ws = wb[SHEET]

        def num(addr):
            v = ws[addr].value
            try:
                return float(v)
            except (TypeError, ValueError):
                return 0.0

        data = {
            "ticker": str(ws[cells["ticker"]].value or "").strip().upper(),
            "revenue": num(cells["revenue"]),
            "netIncome": num(cells["netIncome"]),
            "eps": num(cells["eps"]),
            "shares": num(cells["shares"]),
            "workbook": WORKBOOK,
        }
        wb.close()
        return data
    except Exception as exc:                                     # noqa: BLE001
        return {"error": f"Could not read workbook: {exc}"}


def read_api_key():
    """Pull the Alpha Vantage key out of the workbook, if the user stored it there."""
    try:
        wb_f = openpyxl.load_workbook(WORKBOOK)
        cells = resolve_cells(wb_f)
        wb_f.close()
        wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
        val = wb[SHEET][cells["apiKey"]].value
        wb.close()
        key = str(val or "").strip()
        return "" if key.upper().startswith("YOUR_API_KEY") else key
    except Exception:                                            # noqa: BLE001
        return ""


# ── path 1: real Excel Power Query refresh via COM ──────────────────────────
def refresh_via_excel(ticker, api_key):
    """Drive Excel itself: set the ticker, RefreshAll (Power Query), save, read back.

    Returns a dict on success, or raises with a readable reason.
    """
    try:
        import pythoncom  # noqa: F401  (pywin32)
        import win32com.client as win32
    except ImportError as exc:
        raise RuntimeError("pywin32 not installed") from exc

    import pythoncom
    pythoncom.CoInitialize()
    excel = None
    opened_here = False
    try:
        try:
            excel = win32.GetActiveObject("Excel.Application")
        except Exception:                                        # noqa: BLE001
            excel = win32.DispatchEx("Excel.Application")
            opened_here = True
        excel.DisplayAlerts = False

        target = os.path.normcase(os.path.abspath(WORKBOOK))
        wb = None
        for i in range(1, excel.Workbooks.Count + 1):
            cand = excel.Workbooks.Item(i)
            try:
                if os.path.normcase(os.path.abspath(cand.FullName)) == target:
                    wb = cand
                    break
            except Exception:                                    # noqa: BLE001
                continue
        if wb is None:
            wb = excel.Workbooks.Open(os.path.abspath(WORKBOOK))
            opened_here = True

        ws = wb.Worksheets(SHEET)

        # Write the ticker (and key, if supplied) so Power Query picks them up.
        def set_named(name, fallback_addr, value):
            if value in (None, ""):
                return
            try:
                wb.Names(name).RefersToRange.Value = value
            except Exception:                                    # noqa: BLE001
                ws.Range(fallback_addr).Value = value

        set_named("Ticker", FALLBACK_CELLS["ticker"], ticker)
        if api_key:
            set_named("ApiKey", FALLBACK_CELLS["apiKey"], api_key)

        # The actual Power Query refresh.
        wb.RefreshAll()
        try:
            excel.CalculateUntilAsyncQueriesDone()
        except Exception:                                        # noqa: BLE001
            time.sleep(4)
        excel.Calculate()

        def num(name, addr):
            try:
                v = wb.Names(name).RefersToRange.Value
            except Exception:                                    # noqa: BLE001
                v = ws.Range(addr).Value
            try:
                return float(v)
            except (TypeError, ValueError):
                return 0.0

        data = {
            "ticker": str(ws.Range(FALLBACK_CELLS["ticker"]).Value or ticker).strip().upper(),
            "revenue": num("Co_Revenue", FALLBACK_CELLS["revenue"]),
            "netIncome": num("Co_NetIncome", FALLBACK_CELLS["netIncome"]),
            "eps": num("Co_EPS", FALLBACK_CELLS["eps"]),
            "shares": num("Co_Shares", FALLBACK_CELLS["shares"]),
            "source": "excel",
            "message": "Excel Power Query refreshed.",
        }

        try:
            wb.Save()
        except Exception:                                        # noqa: BLE001
            pass
        if opened_here:
            try:
                wb.Close(SaveChanges=True)
                excel.Quit()
            except Exception:                                    # noqa: BLE001
                pass
        return data
    finally:
        pythoncom.CoUninitialize()


# ── path 2: call Alpha Vantage directly (same endpoints as the M code) ──────
def fetch_alpha_vantage(ticker, api_key):
    """Server-side twin of the workbook's M code. No CORS problem from here."""
    if not api_key:
        raise RuntimeError("No Alpha Vantage API key. Enter one in the dashboard "
                           "or in the workbook's API Key cell.")

    def get(fn):
        url = ("https://www.alphavantage.co/query?function=" + fn
               + "&symbol=" + urllib.parse.quote(ticker)
               + "&apikey=" + urllib.parse.quote(api_key))
        req = urllib.request.Request(url, headers={"User-Agent": "ProjectionModel/1.0"})
        with urllib.request.urlopen(req, timeout=25) as resp:
            return json.loads(resp.read().decode("utf-8"))

    overview = get("OVERVIEW")
    income = get("INCOME_STATEMENT")

    # Alpha Vantage signals throttling / bad keys with these keys instead of data.
    for payload in (overview, income):
        for flag in ("Note", "Information", "Error Message"):
            if isinstance(payload, dict) and flag in payload:
                raise RuntimeError(str(payload[flag])[:300])

    reports = (income or {}).get("annualReports") or []
    if not reports:
        raise RuntimeError(f"No annual income statement returned for {ticker}.")
    latest = reports[0]

    def num(value, divisor=1.0):
        try:
            return float(value) / divisor
        except (TypeError, ValueError):
            return 0.0

    return {
        "ticker": ticker,
        "revenue": num(latest.get("totalRevenue"), 1e6),
        "netIncome": num(latest.get("netIncome"), 1e6),
        "eps": num(overview.get("EPS")),
        "shares": num(overview.get("SharesOutstanding"), 1e6),
        "source": "alphavantage",
        "message": "Fetched from Alpha Vantage (Excel not driven).",
    }


def write_into_workbook(data):
    """Best-effort: park fetched values on Data_API so Excel shows them too.

    Silently skips when the workbook is open in Excel (file locked) — the
    dashboard already has the numbers either way.
    """
    try:
        wb = openpyxl.load_workbook(WORKBOOK)
        if "Data_API" not in wb.sheetnames:
            wb.close()
            return False
        ws = wb["Data_API"]
        target = None
        for r in range(2, max(3, ws.max_row + 2)):
            cell = ws.cell(row=r, column=1).value
            if cell is None or str(cell).strip().upper() == data["ticker"]:
                target = r
                break
        if target is None:
            target = ws.max_row + 1
        ws.cell(row=target, column=1, value=data["ticker"])
        ws.cell(row=target, column=2, value=data["revenue"])
        ws.cell(row=target, column=3, value=data["netIncome"])
        ws.cell(row=target, column=4, value=data["eps"])
        ws.cell(row=target, column=5, value=data["shares"])
        wb.save(WORKBOOK)
        wb.close()
        return True
    except Exception:                                            # noqa: BLE001
        return False


# ── refresh orchestration ───────────────────────────────────────────────────
def refresh_chain(ticker, api_key):
    """Try each refresh path in order. Returns (payload, http_status)."""
    notes = []

    # 1. real Power Query refresh, driven through Excel itself
    try:
        data = refresh_via_excel(ticker, api_key)
        if any(data[k] for k in ("revenue", "netIncome", "eps", "shares")):
            return data, 200
        notes.append("Excel refreshed but Company Inputs are still empty "
                     "(the Power Query is not connected in the saved workbook).")
    except Exception as exc:                                     # noqa: BLE001
        notes.append(f"Excel COM path unavailable: {exc}")

    # 2. call the same endpoints the M code uses, from here
    try:
        data = fetch_alpha_vantage(ticker, api_key)
        wrote = write_into_workbook(data)
        data["message"] = ("Fetched from Alpha Vantage"
                           + (" and written to Data_API." if wrote
                              else " (workbook is open, so it was not written)."))
        if notes:
            data["note"] = " ".join(notes)
        return data, 200
    except Exception as exc:                                     # noqa: BLE001
        notes.append(f"Alpha Vantage path failed: {exc}")

    # 3. fall back to whatever the workbook already holds
    cached = read_company()
    if "error" not in cached and any(
        cached[k] for k in ("revenue", "netIncome", "eps", "shares")
    ):
        cached["source"] = "workbook-cache"
        cached["message"] = "Showing the workbook's last saved values."
        cached["note"] = " ".join(notes)
        return cached, 200

    return {"error": " ".join(notes) or "Refresh failed."}, 502


# ── Notion Research: turn a posted projection into Notion blocks ────────────
# All the actual math (revenue/NI/EPS/CAGR/growth, etc.) happens client-side
# in dashboard.html; this only formats numbers it already computed into
# Notion's block JSON. Scenario shape mirrors computeScenario() there:
# revenue/netIncome/eps/priceLow/priceHigh arrays (one entry per year), plus
# cagrLow/cagrHigh and the revGrowth/niGrowth/peLow/peHigh assumption arrays.
#
# v1 deliberately has no "current price" field — nothing in this app fetches
# a live quote today (Alpha Vantage is called here for fundamentals only), so
# rather than fabricate a number this is left out until that's added.
SCENARIO_LABELS = {"base": "Base Case", "bear": "Bear Case", "bull": "Bull Case"}


def _fmt_money(v):
    try:
        v = float(v)
    except (TypeError, ValueError):
        return "—"
    return f"{'-' if v < 0 else ''}${abs(v):,.0f}"


def _fmt_dollars(v):
    try:
        return f"${float(v):.2f}"
    except (TypeError, ValueError):
        return "—"


def _fmt_pct(v):
    try:
        return f"{float(v) * 100:.1f}%"
    except (TypeError, ValueError):
        return "—"


def _fmt_x(v):
    try:
        return f"{float(v):.1f}x"
    except (TypeError, ValueError):
        return "—"


def _summarize_series(arr, fmt, suffix=""):
    """A flat value shows as one number; a value that varies by year says so
    instead of quietly picking one year and implying it's the whole story."""
    vals = [v for v in (arr or []) if isinstance(v, (int, float))]
    if not vals:
        return "—"
    lo, hi = min(vals), max(vals)
    if abs(hi - lo) < 1e-9:
        return f"{fmt(lo)}{suffix}"
    return f"{fmt(lo)}–{fmt(hi)}{suffix} (varies by year)"


def _table_block(headers, rows):
    def cell(text):
        return [{"type": "text", "text": {"content": str(text)[:180]}}]

    table_rows = [{"object": "block", "type": "table_row",
                   "table_row": {"cells": [cell(h) for h in headers]}}]
    for row in rows:
        table_rows.append({"object": "block", "type": "table_row",
                            "table_row": {"cells": [cell(c) for c in row]}})
    return {
        "object": "block", "type": "table",
        "table": {
            "table_width": len(headers),
            "has_column_header": True,
            "has_row_header": True,
            "children": table_rows,
        },
    }


def build_projection_blocks(payload):
    """payload: {"years": [...], "scenarios": {"base": {...}, "bear": {...}, "bull": {...}}}"""
    years = payload.get("years") or []
    scenarios = payload.get("scenarios") or {}
    blocks = []

    val_rows = []
    for key in ("base", "bear", "bull"):
        sc = scenarios.get(key) or {}
        price_low = (sc.get("priceLow") or [None])[-1]
        price_high = (sc.get("priceHigh") or [None])[-1]
        val_rows.append([
            SCENARIO_LABELS[key],
            f"{_fmt_pct(sc.get('cagrLow'))} – {_fmt_pct(sc.get('cagrHigh'))}",
            f"{_fmt_dollars(price_low)} – {_fmt_dollars(price_high)}",
        ])
    blocks.append(notion_service.heading_block("Implied Valuation"))
    blocks.append(_table_block(
        ["Scenario", "CAGR", f"{years[-1]} Price" if years else "Final-Year Price"], val_rows))

    for key in ("base", "bear", "bull"):
        sc = scenarios.get(key)
        if not sc:
            continue
        blocks.append(notion_service.heading_block(SCENARIO_LABELS[key]))
        headers = ["Metric"] + [str(y) for y in years]
        rows = [
            ["Revenue ($M)"] + [_fmt_money(v) for v in sc.get("revenue", [])],
            ["Net Income ($M)"] + [_fmt_money(v) for v in sc.get("netIncome", [])],
            ["EPS ($/sh)"] + [_fmt_dollars(v) for v in sc.get("eps", [])],
            ["Share Price Low"] + [_fmt_dollars(v) for v in sc.get("priceLow", [])],
            ["Share Price High"] + [_fmt_dollars(v) for v in sc.get("priceHigh", [])],
        ]
        blocks.append(_table_block(headers, rows))

        assumptions_line = (
            f"Revenue growth {_summarize_series(sc.get('revGrowth', [])[1:], _fmt_pct)} · "
            f"Net income growth {_summarize_series(sc.get('niGrowth', [])[1:], _fmt_pct)} · "
            f"P/E {_summarize_series(sc.get('peLow', []), _fmt_x)}"
            f"–{_summarize_series(sc.get('peHigh', []), _fmt_x)}"
        )
        blocks.append(notion_service.paragraph_block(assumptions_line))

    return blocks


# ── HTTP layer ──────────────────────────────────────────────────────────────
class Handler(BaseHTTPRequestHandler):
    server_version = "ProjectionBridge/1.0"

    def log_message(self, fmt, *args):
        sys.stderr.write("  %s\n" % (fmt % args))

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, DELETE, OPTIONS")

    def _json(self, obj, status=200):
        body = json.dumps(obj).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self._cors()
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):                                        # noqa: N802
        # Preflight requests never carry credentials — don't gate them, or
        # browsers can't even ask permission to send the real request.
        self.send_response(204)
        self._cors()
        self.end_headers()

    def do_HEAD(self):                                           # noqa: N802
        # Cloud platforms probe liveness with HEAD, with no credentials —
        # answer plainly so a health check never reads as "service down"
        # just because it wasn't logged in. Carries no body either way.
        self.send_response(200)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", "0")
        self.end_headers()

    def _authorized(self):
        if not AUTH_ENABLED:
            return True
        header = self.headers.get("Authorization", "")
        if not header.startswith("Basic "):
            return False
        # constant-time compare — this endpoint is reachable from the public
        # internet once tunneled, so a timing side-channel is a real risk.
        return hmac.compare_digest(header[6:], AUTH_TOKEN)

    def _require_auth(self):
        body = b"Authentication required."
        self.send_response(401)
        self.send_header("WWW-Authenticate", 'Basic realm="Projection Model"')
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self._cors()
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):                                            # noqa: N802
        if not self._authorized():
            self._require_auth()
            return
        path = urllib.parse.urlparse(self.path).path

        if path in ("/", "/index.html", "/dashboard.html"):
            fp = os.path.join(HERE, "dashboard.html")
            try:
                with open(fp, "rb") as fh:
                    body = fh.read()
            except OSError:
                self._json({"error": "dashboard.html missing"}, 404)
                return
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(body)
            return

        # PWA assets — manifest, service worker, icons. Static, cacheable, and
        # what makes the browser offer a real "Install app" prompt.
        STATIC_FILES = {
            "/manifest.json": ("manifest.json", "application/manifest+json"),
            "/sw.js": ("sw.js", "application/javascript"),
        }
        if path in STATIC_FILES:
            rel, ctype = STATIC_FILES[path]
            self._serve_static(os.path.join(HERE, rel), ctype)
            return

        if path.startswith("/icons/"):
            # Path-traversal guard: resolve, then require it stay under icons/.
            icons_dir = os.path.realpath(os.path.join(HERE, "icons"))
            requested = os.path.realpath(os.path.join(HERE, path.lstrip("/")))
            if os.path.commonpath([icons_dir, requested]) == icons_dir and os.path.isfile(requested):
                self._serve_static(requested, "image/png")
            else:
                self._json({"error": "not found"}, 404)
            return

        if path == "/api/health":
            self._json({"ok": True, "workbook": WORKBOOK,
                        "workbookExists": os.path.exists(WORKBOOK)})
            return

        if path == "/api/model":
            data = read_company()
            self._json(data, 500 if "error" in data else 200)
            return

        if path == "/api/projects":
            self._json(list_saved_projects())
            return

        if path == "/api/notion/status":
            if not NOTION_TOKEN:
                self._json({"configured": False, "connected": False, "workspace": None,
                            "database": get_notion_settings(), "error": None})
                return
            ok, info = notion_service.check_token(NOTION_TOKEN)
            self._json({
                "configured": True,
                "connected": ok,
                "workspace": info.get("name") if ok else None,
                "database": get_notion_settings(),
                "error": None if ok else notion_service.error_message(info),
            })
            return

        if path == "/api/watchlist":
            tickers = get_watchlist()
            self._json({"tickers": tickers, "quotes": quotes_service.fetch_quotes(tickers)})
            return

        if path == "/api/watchlist/history":
            qs = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
            ticker = (qs.get("ticker") or [""])[0].strip().upper()
            if not TICKER_RE.match(ticker):
                self._json({"error": "Invalid ticker."}, 400)
                return
            points, err = quotes_service.fetch_history(ticker)
            if err:
                self._json({"ticker": ticker, "points": [], "error": err}, 502)
                return
            self._json({"ticker": ticker, "points": points, "error": None})
            return

        if path == "/api/notion/databases":
            if not NOTION_TOKEN:
                self._json({"error": "Notion isn't configured on the server yet."}, 400)
                return
            ok, data = notion_service.list_databases(NOTION_TOKEN)
            if not ok:
                self._notion_fail(data)
                return
            self._json({"databases": data})
            return

        self._json({"error": "not found"}, 404)

    def _notion_fail(self, err):
        status = err.get("status") if isinstance(err, dict) else 0
        # Pass real Notion HTTP statuses through as-is; anything else (a
        # network failure, a status Notion doesn't normally send here) becomes
        # a 502 — this bridge's fault for not reaching Notion, not the caller's.
        http_status = status if status in (400, 401, 403, 404, 429) else 502
        self._json({"error": notion_service.error_message(err)}, http_status)

    def _serve_static(self, filepath, content_type):
        try:
            with open(filepath, "rb") as fh:
                body = fh.read()
        except OSError:
            self._json({"error": "not found"}, 404)
            return
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "public, max-age=3600")
        self._cors()
        self.end_headers()
        self.wfile.write(body)

    def _read_json_body(self):
        length = int(self.headers.get("Content-Length") or 0)
        try:
            return json.loads(self.rfile.read(length) or b"{}")
        except json.JSONDecodeError:
            return {}

    def do_POST(self):                                           # noqa: N802
        if not self._authorized():
            self._require_auth()
            return
        path = urllib.parse.urlparse(self.path).path

        if path == "/api/refresh":
            payload = self._read_json_body()
            ticker = str(payload.get("ticker") or "").strip().upper()
            api_key = str(payload.get("apiKey") or "").strip() or read_api_key()
            if not ticker:
                ticker = (read_company().get("ticker") or "AAPL")
            data, status = refresh_chain(ticker, api_key)
            self._json(data, status)
            return

        if path == "/api/projects":
            payload = self._read_json_body()
            ok, err = write_saved_project(payload)
            if not ok:
                self._json({"error": err}, 400)
                return
            self._json({"ok": True})
            return

        if path == "/api/watchlist":
            payload = self._read_json_body()
            raw = payload.get("tickers")
            if not isinstance(raw, list):
                self._json({"error": "tickers must be a list."}, 400)
                return
            ok, result = set_watchlist(raw)
            if not ok:
                self._json({"error": result}, 500)
                return
            self._json({"tickers": result, "quotes": quotes_service.fetch_quotes(result)})
            return

        if path == "/api/notion/settings":
            if not NOTION_TOKEN:
                self._json({"error": "Notion isn't configured on the server yet."}, 400)
                return
            payload = self._read_json_body()
            db_id = str(payload.get("databaseId") or "").strip()
            db_name = str(payload.get("databaseName") or "").strip()
            if not db_id or not db_name:
                self._json({"error": "Pick a database first."}, 400)
                return
            ok, err = set_notion_settings(db_id, db_name)
            if not ok:
                self._json({"error": err}, 500)
                return
            self._json({"ok": True})
            return

        if path == "/api/notion/sync":
            self._handle_notion_sync()
            return

        if path == "/api/notion/notes":
            if not NOTION_TOKEN:
                self._json({"error": "Notion isn't configured on the server yet."}, 400)
                return
            payload = self._read_json_body()
            notion_state = payload.get("notion") or {}
            page_id = notion_state.get("pageId")
            if not page_id:
                self._json({"error": "Sync to Notion once first — after that, notes keep themselves in sync."}, 400)
                return
            ok, new_ids = notion_service.replace_notes_zone(
                NOTION_TOKEN, page_id, notion_state.get("notesHeadingId"),
                notion_state.get("notesBlockIds") or [], str(payload.get("notes") or ""))
            if not ok:
                self._notion_fail(new_ids)
                return
            self._json({"ok": True, "notesBlockIds": new_ids, "lastSyncedAt": int(time.time() * 1000)})
            return

        self._json({"error": "not found"}, 404)

    def _handle_notion_sync(self):
        if not NOTION_TOKEN:
            self._json({"error": "Notion isn't configured on the server yet."}, 400)
            return
        settings = get_notion_settings()
        if not settings:
            self._json({"error": "Pick a Notion database first, in Notion Research settings."}, 400)
            return
        payload = self._read_json_body()
        ticker = str(payload.get("ticker") or "").strip().upper()
        if not ticker:
            self._json({"error": "No ticker to sync — enter Company Inputs first."}, 400)
            return
        notes = str(payload.get("notes") or "")
        existing = payload.get("notion") or {}
        database_id = settings["databaseId"]

        ok, title_prop = notion_service.get_database_title_property(NOTION_TOKEN, database_id)
        if not ok:
            self._notion_fail(title_prop)
            return

        data_blocks = build_projection_blocks(payload)
        synced_label = time.strftime("%b %d, %Y %H:%M UTC", time.gmtime())

        page_id = existing.get("pageId")
        # Only trust existing.notesBlockId/dataBlockIds when page_id itself
        # is confirmed valid below — they were captured against that exact
        # page, and mean nothing (or worse, point at some other page's
        # blocks) the moment we have to fall back to a different page.
        page_verified = False
        if page_id:
            # A page id we tracked before doesn't mean it still exists — the
            # user (or someone else in the workspace) may have deleted it.
            ok, _ = notion_service.get_page(NOTION_TOKEN, page_id)
            page_verified = ok
            if not ok:
                page_id = None

        if not page_id:
            # Might already exist from an earlier sync this record didn't
            # track, or a page made by hand — reuse it rather than create a
            # second page for the same ticker.
            ok, found = notion_service.find_page_by_ticker(NOTION_TOKEN, database_id, title_prop, ticker)
            if not ok:
                self._notion_fail(found)
                return
            page_id = found["id"] if found else None

        if not page_id:
            ok, result = notion_service.create_stock_page(
                NOTION_TOKEN, database_id, title_prop, ticker, notes, data_blocks, synced_label)
            if not ok:
                self._notion_fail(result)
                return
            result["lastSyncedAt"] = int(time.time() * 1000)
            self._json({"ok": True, "notion": result})
            return

        notes_heading_id = existing.get("notesHeadingId") if page_verified else None
        notes_block_ids = (existing.get("notesBlockIds") or []) if page_verified else []
        if not notes_heading_id:
            ok, notes_zone = notion_service.append_notes_zone(NOTION_TOKEN, page_id, notes)
            if not ok:
                self._notion_fail(notes_zone)
                return
            notes_heading_id = notes_zone["notesHeadingId"]
            notes_block_ids = notes_zone["notesBlockIds"]

        old_data_ids = (existing.get("dataBlockIds") or []) if page_verified else []
        ok, new_ids = notion_service.replace_data_section(
            NOTION_TOKEN, page_id, old_data_ids, data_blocks, synced_label)
        if not ok:
            self._notion_fail(new_ids)
            return
        ok, page = notion_service.get_page(NOTION_TOKEN, page_id)
        url = page.get("url") if ok else f"https://notion.so/{page_id.replace('-', '')}"
        self._json({"ok": True, "notion": {
            "pageId": page_id,
            "url": url,
            "notesHeadingId": notes_heading_id,
            "notesBlockIds": notes_block_ids,
            "dataBlockIds": new_ids,
            "lastSyncedAt": int(time.time() * 1000),
        }})

    def do_DELETE(self):                                         # noqa: N802
        if not self._authorized():
            self._require_auth()
            return
        path = urllib.parse.urlparse(self.path).path
        prefix = "/api/projects/"
        if not path.startswith(prefix):
            self._json({"error": "not found"}, 404)
            return
        project_id = path[len(prefix):]
        ok = delete_saved_project(project_id)
        self._json({"ok": ok})


def lan_ip():
    """Best-effort local network address, for opening the app on a phone."""
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))       # no packets sent; just picks the route
        return s.getsockname()[0]
    except Exception:                                            # noqa: BLE001
        return None
    finally:
        s.close()


def main():
    # --lan (or --host 0.0.0.0) exposes the app to your local network so a
    # phone on the same Wi-Fi can open it. Off by default: loopback only.
    # On a cloud host (Render etc.) $PORT is set and there's no LAN concept —
    # bind 0.0.0.0 automatically, since that's the only way it's reachable.
    args = sys.argv[1:]
    host = "0.0.0.0" if IS_CLOUD else "127.0.0.1"
    if "--lan" in args:
        host = "0.0.0.0"
    if "--host" in args:
        try:
            host = args[args.index("--host") + 1]
        except IndexError:
            pass

    print("=" * 68)
    print("  Projection Model — Excel bridge")
    print("=" * 68)
    if IS_CLOUD:
        print("  mode     : cloud (no Excel available — Alpha Vantage refresh only)")
    print(f"  workbook : {WORKBOOK}")
    print(f"  exists   : {os.path.exists(WORKBOOK)}")
    if AUTH_ENABLED:
        print(f"  login    : {AUTH_USER} / {AUTH_PASS}")
        if not (os.environ.get("BRIDGE_USERNAME") and os.environ.get("BRIDGE_PASSWORD")):
            print(f"             (saved in .bridge_credentials.json — delete it for a new password)")
    else:
        print("  login    : DISABLED (--no-auth) — do not expose this beyond localhost")
    try:
        import win32com.client  # noqa: F401
        print("  pywin32  : available (Excel Power Query refresh enabled)")
    except ImportError:
        print("  pywin32  : NOT installed — refresh will use Alpha Vantage directly.")
        if not IS_CLOUD:
            print("             enable the Excel path with:  pip install pywin32")

    if USE_PG:
        ok, err = pg_init()
        if ok:
            print("  saves    : Postgres (NEON_DATABASE_URL) — synced across devices")
        else:
            print("  saves    : Postgres CONFIGURED BUT UNREACHABLE — saving will fail")
            print(f"             {err}")
    elif IS_CLOUD:
        print("  saves    : local files — WIPED ON EVERY RESTART on this host.")
        print("             set NEON_DATABASE_URL to keep them and sync across devices.")
    else:
        print(f"  saves    : {SAVES_DIR}")

    if NOTION_TOKEN:
        settings = get_notion_settings()
        print("  notion   : token set" + (f" — syncing to \"{settings['databaseName']}\"" if settings else " — no database selected yet"))
    else:
        print("  notion   : not configured — set NOTION_TOKEN to enable Notion Research")

    url = f"http://127.0.0.1:{PORT}/"
    print(f"  serving  : {url}" if not IS_CLOUD else f"  serving  : 0.0.0.0:{PORT}")
    if IS_CLOUD:
        pass
    elif host == "0.0.0.0":
        ip = lan_ip()
        if ip:
            print(f"  on phone : http://{ip}:{PORT}/   (same Wi-Fi)")
            print("             Windows Firewall may ask to allow Python — say yes,")
            print("             and make sure it's allowed on Private networks.")
        else:
            print("  on phone : could not determine this machine's LAN address")
    else:
        print(f"  phone    : restart with  --lan  to allow access from your phone")
    print("  press Ctrl+C to stop")
    print("=" * 68)
    # Cloud log pipes aren't a TTY, so stdout is fully block-buffered by
    # default — without this, the whole banner above (including the login
    # line) can sit unflushed indefinitely on a long-running server and
    # never actually reach the platform's log viewer.
    sys.stdout.flush()

    httpd = ThreadingHTTPServer((host, PORT), Handler)
    if "--no-open" not in args and not IS_CLOUD:
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n  stopped.")
        httpd.server_close()


if __name__ == "__main__":
    main()
